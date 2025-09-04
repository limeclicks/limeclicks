"""
Consolidate multiple CSV audit reports into organized Excel files with tabs
"""

import os
import io
import csv
import logging
from pathlib import Path
from typing import Dict, List, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class AuditReportConsolidator:
    """Consolidates multiple CSV files into organized Excel workbooks"""
    
    # Define report groups and their included file patterns
    REPORT_GROUPS = {
        'Technical_SEO_Audit': {
            'display_name': 'Technical SEO Audit',
            'description': 'Core technical SEO issues and crawl data',
            'tabs': {
                'Crawl Overview': ['crawl_overview'],
                'Issues Summary': ['issues_overview', 'issues_reports'],
                'Response Codes': ['response_codes', 'client_error', 'server_error', 
                                  'redirection', 'success', 'no_response'],
                'Indexability': ['internal_all', 'indexable', 'non_indexable'],
                'External URLs': ['external_all'],
            }
        },
        'Content_Optimization': {
            'display_name': 'Content Optimization Report',
            'description': 'Page titles, meta descriptions, and content analysis',
            'tabs': {
                'Page Titles': ['page_titles', 'page_title', 'titles', 
                              'title_too_long', 'title_too_short', 'missing_title',
                              'duplicate_title', 'title_same_h1'],
                'Meta Descriptions': ['meta_description', 'meta_desc', 
                                    'meta_too_long', 'meta_too_short',
                                    'missing_meta', 'duplicate_meta'],
                'H1 Tags': ['h1_all', 'h1_1', 'h1-all', 'missing_h1', 
                          'duplicate_h1', 'multiple_h1'],
                'H2 Tags': ['h2_all', 'h2_1', 'h2-all', 'missing_h2', 
                          'duplicate_h2'],
                'Content Issues': ['low_content', 'thin_content', 'duplicate_content',
                                 'readability', 'spelling', 'grammar']
            }
        },
        'Technical_Configuration': {
            'display_name': 'Technical Configuration',
            'description': 'Canonicals, directives, structured data, and technical setup',
            'tabs': {
                'Canonical URLs': ['canonical', 'canonicals', 'canonical_chains',
                                 'canonicalised', 'non_canonical'],
                'Robots & Directives': ['directives', 'robots', 'robots_txt',
                                      'meta_robots', 'x_robots'],
                'Structured Data': ['structured_data', 'schema', 'json_ld',
                                  'microdata', 'rdfa'],
                'Hreflang': ['hreflang', 'hreflang_issues'],
                'Sitemaps': ['sitemap', 'xml_sitemap', 'sitemap_issues']
            }
        },
        'Media_Analysis': {
            'display_name': 'Media & Resources Analysis',
            'description': 'Images, scripts, stylesheets, and other resources',
            'tabs': {
                'Images': ['images', 'images_missing_alt', 'images_over',
                         'broken_images', 'image_issues'],
                'JavaScript': ['javascript', 'js', 'scripts', 'blocked_js'],
                'CSS': ['css', 'stylesheets', 'blocked_css'],
                'Other Resources': ['fonts', 'videos', 'audio', 'flash']
            }
        },
        'Link_Analysis': {
            'display_name': 'Link Analysis',
            'description': 'Internal and external link analysis',
            'tabs': {
                'All Links': ['links', 'all_links'],
                'Internal Links': ['inlinks', 'internal_links', 'internal_outlinks'],
                'External Links': ['outlinks', 'external_links', 'external_outlinks'],
                'Broken Links': ['broken_links', 'link_errors'],
                'Redirect Chains': ['redirect_chains', 'redirect_loops']
            }
        },
        'Page_Performance': {
            'display_name': 'Page Performance',
            'description': 'Page speed, Core Web Vitals, and performance metrics',
            'tabs': {
                'Page Speed': ['page_speed', 'load_time', 'response_time'],
                'Core Web Vitals': ['core_web_vitals', 'lcp', 'fid', 'cls'],
                'Page Size': ['page_size', 'page_weight'],
                'Performance Issues': ['performance_issues', 'slow_pages']
            }
        },
        'Security_Issues': {
            'display_name': 'Security & HTTPS',
            'description': 'Security issues and HTTPS configuration',
            'tabs': {
                'HTTPS Issues': ['https', 'mixed_content', 'insecure_content'],
                'Security Headers': ['security_headers', 'hsts', 'csp', 'x_frame'],
                'Certificates': ['ssl_certificate', 'certificate_issues']
            }
        },
        'Validation_Other': {
            'display_name': 'Validation & Other Reports',
            'description': 'HTML validation and miscellaneous reports',
            'tabs': {
                'HTML Validation': ['validation', 'html_validation', 'w3c_errors'],
                'AMP Issues': ['amp', 'amp_validation'],
                'Pagination': ['pagination', 'rel_next_prev'],
                'Custom Extraction': ['custom', 'extraction', 'xpath'],
                'Other': ['other', 'uncategorized']
            }
        }
    }
    
    def __init__(self, audit_dir: str):
        """
        Initialize the consolidator with an audit directory
        
        Args:
            audit_dir: Path to directory containing CSV files
        """
        self.audit_dir = Path(audit_dir)
        if not self.audit_dir.exists():
            raise ValueError(f"Audit directory not found: {audit_dir}")
        
        # Find all CSV files
        self.csv_files = list(self.audit_dir.glob('*.csv'))
        logger.info(f"Found {len(self.csv_files)} CSV files to consolidate")
        
    def consolidate_reports(self) -> Dict[str, bytes]:
        """
        Consolidate CSV files into grouped Excel workbooks
        
        Returns:
            Dict mapping report names to Excel file bytes
        """
        consolidated_reports = {}
        
        for group_key, group_config in self.REPORT_GROUPS.items():
            excel_data = self._create_excel_workbook(group_config)
            
            if excel_data:
                # Generate filename
                filename = f"{group_config['display_name'].replace(' ', '_')}.xlsx"
                consolidated_reports[filename] = excel_data
                logger.info(f"Created consolidated report: {filename}")
        
        # Also create a master summary report
        summary_data = self._create_summary_report()
        if summary_data:
            consolidated_reports['Audit_Summary.xlsx'] = summary_data
            
        return consolidated_reports
    
    def _create_excel_workbook(self, group_config: Dict) -> bytes:
        """
        Create an Excel workbook for a report group
        
        Args:
            group_config: Configuration for the report group
            
        Returns:
            Excel file as bytes, or None if no data
        """
        try:
            # Create workbook
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove default sheet
            has_data = False
            
            # Process each tab
            for tab_name, file_patterns in group_config['tabs'].items():
                df = self._combine_csv_files(file_patterns)
                
                if df is not None and not df.empty:
                    # Create worksheet
                    ws = workbook.create_sheet(title=self._clean_sheet_name(tab_name))
                    
                    # Add data
                    for row in dataframe_to_rows(df, index=False, header=True):
                        ws.append(row)
                    
                    # Format header row
                    self._format_excel_sheet(ws)
                    has_data = True
            
            if not has_data:
                return None
            
            # Add summary sheet at the beginning
            summary_ws = workbook.create_sheet(title="Summary", index=0)
            self._add_summary_sheet(summary_ws, group_config)
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.read()
            
        except Exception as e:
            logger.error(f"Error creating Excel workbook: {e}")
            return None
    
    def _combine_csv_files(self, patterns: List[str]) -> pd.DataFrame:
        """
        Combine multiple CSV files matching the patterns into a single DataFrame
        
        Args:
            patterns: List of filename patterns to match
            
        Returns:
            Combined DataFrame or None if no matching files
        """
        combined_df = None
        
        for csv_file in self.csv_files:
            filename_lower = csv_file.name.lower()
            
            # Check if file matches any pattern
            matches = False
            for pattern in patterns:
                if pattern.lower() in filename_lower:
                    matches = True
                    break
            
            if matches:
                try:
                    # Read CSV file
                    df = pd.read_csv(csv_file, encoding='utf-8', low_memory=False)
                    
                    # Add source filename column
                    df['Source_File'] = csv_file.name
                    
                    # Combine dataframes
                    if combined_df is None:
                        combined_df = df
                    else:
                        # Concatenate with matching columns
                        combined_df = pd.concat([combined_df, df], ignore_index=True, sort=False)
                        
                except Exception as e:
                    logger.warning(f"Error reading CSV file {csv_file.name}: {e}")
                    continue
        
        return combined_df
    
    def _format_excel_sheet(self, worksheet):
        """Apply formatting to Excel worksheet"""
        # Format header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:  # First row
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_summary_sheet(self, worksheet, group_config):
        """Add a summary sheet to the workbook"""
        # Add title
        worksheet['A1'] = group_config['display_name']
        worksheet['A1'].font = Font(bold=True, size=14)
        
        worksheet['A3'] = 'Description:'
        worksheet['B3'] = group_config['description']
        
        worksheet['A5'] = 'Report Contents:'
        worksheet['A5'].font = Font(bold=True)
        
        # List tabs
        row = 6
        for tab_name in group_config['tabs'].keys():
            worksheet[f'A{row}'] = f"• {tab_name}"
            row += 1
        
        # Add generation info
        row += 1
        worksheet[f'A{row}'] = 'Generated:'
        worksheet[f'B{row}'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Format columns
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 60
    
    def _create_summary_report(self) -> bytes:
        """
        Create a master summary report with key metrics from all reports
        
        Returns:
            Excel file as bytes
        """
        try:
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Audit Summary"
            
            # Add summary information
            ws['A1'] = 'Site Audit Summary Report'
            ws['A1'].font = Font(bold=True, size=16)
            
            # Count files by type
            ws['A3'] = 'Files Processed'
            ws['A3'].font = Font(bold=True)
            
            file_counts = {}
            for csv_file in self.csv_files:
                base_name = csv_file.stem.split('_')[0]  # Get first part of filename
                file_counts[base_name] = file_counts.get(base_name, 0) + 1
            
            row = 4
            for file_type, count in sorted(file_counts.items()):
                ws[f'A{row}'] = file_type
                ws[f'B{row}'] = count
                row += 1
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.read()
            
        except Exception as e:
            logger.error(f"Error creating summary report: {e}")
            return None
    
    def _clean_sheet_name(self, name: str) -> str:
        """Clean sheet name to be Excel-compatible (max 31 chars, no special chars)"""
        # Remove invalid characters
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '')
        
        # Truncate to 31 characters (Excel limit)
        return name[:31]
    
    def get_file_mapping(self) -> Dict[str, List[str]]:
        """
        Get mapping of which CSV files go into which consolidated report
        
        Returns:
            Dict mapping consolidated report name to list of included CSV files
        """
        mapping = {}
        
        for group_key, group_config in self.REPORT_GROUPS.items():
            report_name = f"{group_config['display_name'].replace(' ', '_')}.xlsx"
            included_files = []
            
            for tab_name, patterns in group_config['tabs'].items():
                for csv_file in self.csv_files:
                    filename_lower = csv_file.name.lower()
                    for pattern in patterns:
                        if pattern.lower() in filename_lower:
                            included_files.append(csv_file.name)
                            break
            
            if included_files:
                mapping[report_name] = list(set(included_files))  # Remove duplicates
        
        return mapping